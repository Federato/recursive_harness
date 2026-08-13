"""Phase 3: extract the remaining package categories - DOC/*.xlsx,
STC/*.json, Form Fields, Form Pages, Form Related Fields, Ratebook
Columns, Ratebook Tables.

Emits:
  out/doc_sheets.csv     sheet inventory of every DOC workbook: package,
                         workbook name, sheet name, dimensions, header row
  out/doc_exceptions.csv full extraction of the non-empty data rows of the
                         "Refer to Company", "Not Supported" and "Special
                         Consideration" sheets - i.e. the register of what
                         ERC declines to rate automatically
  out/stc_index.csv      per STC json: package, filename, bytes, top-level
                         keys, the SchemeKeys (ProductName / EffectiveDate),
                         and the number of nested objects and leaf fields
  out/form_csv_schema.csv per (category, file, column) header inventory
                         across the corpus with the packages carrying it
  out/form_pages.csv     full extraction of Form Pages rows (page tree:
                         TableName, Type, Name, ParentName, AttachmentType,
                         form Number, Condition, Status)
  out/doc_stc_report.txt totals, sheet-name vocabulary, STC effective-date
                         vs directory-edition reconciliation, Status code
                         vocabulary in the form CSVs.
"""
from __future__ import annotations

import csv
import json
import sys
from collections import Counter, defaultdict
from multiprocessing import Pool
from pathlib import Path
from importlib import import_module

sys.path.insert(0, str(Path(__file__).parent))
c = import_module("00_common")

EXC_SHEETS = {"Refer to Company", "Not Supported", "Special Consideration"}
FORM_CATS = ["Form Fields", "Form Pages", "Form Related Fields",
             "Ratebook Columns", "Ratebook Tables"]


def count_json(o):
    objs = leaves = 0
    stack = [o]
    while stack:
        x = stack.pop()
        if isinstance(x, dict):
            objs += 1
            stack.extend(x.values())
        elif isinstance(x, list):
            stack.extend(x)
        else:
            leaves += 1
    return objs, leaves


def scan(a):
    pkg_id, juris, edition, content = a
    content = Path(content)
    docs, excs, stcs, schema, pages = [], [], [], [], []
    probs = []

    d = content / "DOC"
    if d.is_dir():
        import openpyxl
        for f in sorted(d.iterdir()):
            if f.suffix.lower() != ".xlsx":
                continue
            try:
                wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            except Exception as e:
                probs.append(f"XLSXFAIL {pkg_id} {f.name}: {e}")
                continue
            for sn in wb.sheetnames:
                ws = wb[sn]
                rows = list(ws.iter_rows(values_only=True))
                hdr = [str(x).strip() if x is not None else ""
                       for x in (rows[0] if rows else [])]
                body = [r for r in rows[1:]
                        if any(x is not None and str(x).strip() for x in r)]
                docs.append((pkg_id, juris, f.name, sn, len(rows),
                             ws.max_column or 0, "|".join(hdr), len(body)))
                if sn in EXC_SHEETS:
                    for r in body:
                        excs.append((pkg_id, juris, edition, sn,
                                     *[str(x).replace("\n", " ").strip()
                                       if x is not None else ""
                                       for x in list(r)[:6]] + [""] * (6 - len(r))))
            wb.close()

    d = content / "STC"
    if d.is_dir():
        for f in sorted(d.iterdir()):
            if f.suffix.lower() != ".json":
                continue
            try:
                o = json.loads(c.read_text(f))
            except Exception as e:
                probs.append(f"JSONFAIL {pkg_id} {f.name}: {e}")
                continue
            sk = o.get("SchemeKeys", {}) if isinstance(o, dict) else {}
            objs, leaves = count_json(o)
            stcs.append((pkg_id, juris, edition, f.name, f.stat().st_size,
                         "|".join(sorted(o.keys())) if isinstance(o, dict) else "",
                         sk.get("ProductName", ""), sk.get("EffectiveDateTime", ""),
                         objs, leaves))

    for cat in FORM_CATS:
        d = content / cat
        if not d.is_dir():
            continue
        for f in sorted(d.iterdir()):
            if f.suffix.lower() != ".csv":
                continue
            try:
                hdr, rdr = c.read_csv_rows(f)
            except Exception as e:
                probs.append(f"CSVFAIL {pkg_id} {f.name}: {e}")
                continue
            n = 0
            for row in rdr:
                n += 1
                if cat == "Form Pages":
                    r = dict(zip(hdr, row))
                    pages.append((pkg_id, juris, edition,
                                  r.get("TableName", ""), r.get("Type", ""),
                                  r.get("Name", ""), r.get("ParentName", ""),
                                  r.get("AttachmentType", ""),
                                  r.get("Number", ""), r.get("Sequence", ""),
                                  r.get("Status", ""),
                                  (r.get("Condition", "") or "")[:400]))
            schema.append((pkg_id, juris, cat, f.name, "|".join(hdr), n))
    return docs, excs, stcs, schema, pages, probs


def main():
    pkgs = c.find_packages()
    args = [(p.pkg_id, p.juris, p.edition, str(p.content)) for p in pkgs]
    D = E = S = C_ = P = []
    D, E, S, C_, P, PR = [], [], [], [], [], []
    with Pool() as pool:
        for d, e, s, sc, pg, pr in pool.imap_unordered(scan, args, chunksize=2):
            D += d; E += e; S += s; C_ += sc; P += pg; PR += pr

    def dump(name, header, rows):
        with open(c.OUT / name, "w", newline="", encoding="utf-8") as fh:
            w = csv.writer(fh); w.writerow(header); w.writerows(rows)

    dump("doc_sheets.csv", ["pkg_id", "juris", "workbook", "sheet", "n_rows",
                            "n_cols", "header", "n_data_rows"], D)
    dump("doc_exceptions.csv", ["pkg_id", "juris", "edition", "sheet",
                                "c1", "c2", "c3", "c4", "c5", "c6"], E)
    dump("stc_index.csv", ["pkg_id", "juris", "edition", "file", "bytes",
                           "top_keys", "product_name", "effective_datetime",
                           "n_objects", "n_leaves"], S)
    dump("form_csv_schema.csv", ["pkg_id", "juris", "category", "file",
                                 "header", "n_rows"], C_)
    dump("form_pages.csv", ["pkg_id", "juris", "edition", "table_name", "type",
                            "name", "parent_name", "attachment_type",
                            "form_number", "sequence", "status", "condition"], P)

    L = []; A = L.append
    A(f"DOC workbooks: {len({(d[0], d[2]) for d in D})}  sheets: {len(D)}")
    A(f"  sheet name vocabulary: {Counter(d[3] for d in D).most_common()}")
    A(f"  workbooks whose sheet set != the standard 5: "
      f"{sum(1 for k, v in Counter((d[0], d[2]) for d in D).items() if v != 5)}")
    A(f"  exception rows extracted: {len(E)}  by sheet: "
      f"{Counter(e[3] for e in E).most_common()}")
    A(f"  packages with >=1 'Refer to Company' row: "
      f"{len({e[0] for e in E if e[3] == 'Refer to Company'})}")
    A(f"  packages with >=1 'Not Supported' row: "
      f"{len({e[0] for e in E if e[3] == 'Not Supported'})}")
    A("")
    A(f"STC files: {len(S)} in {len({s[0] for s in S})} packages")
    A(f"  bytes: {sum(s[4] for s in S)}  leaf fields: {sum(s[9] for s in S)}")
    A(f"  top-level key sets: {Counter(s[5][:120] for s in S).most_common(5)}")
    A(f"  ProductName distinct: {len({s[6] for s in S})}")
    A("  STC EffectiveDateTime vs directory edition:")
    agree = sum(1 for s in S if s[7][:10].replace("-", "") == s[2])
    A(f"    agree: {agree} / {sum(1 for s in S if s[7])} with a date")
    dis = [(s[0], s[2], s[7]) for s in S
           if s[7] and s[7][:10].replace("-", "") != s[2]]
    for x in dis[:15]:
        A(f"    DISAGREE {x[0]} dir_edition={x[1]} stc={x[2]}")
    A("")
    A("FORM / RATEBOOK CSVs")
    for cat in FORM_CATS:
        rs = [r for r in C_ if r[2] == cat]
        hs = Counter(r[4] for r in rs)
        A(f"  {cat}: {len(rs)} files, {sum(r[5] for r in rs)} rows, "
          f"{len(hs)} distinct headers")
        for h, n in hs.most_common(3):
            A(f"     ({n}) {h}")
    A("")
    A(f"Form Pages rows extracted: {len(P)}")
    A(f"  Type vocabulary: {Counter(p[4] for p in P).most_common()}")
    A(f"  AttachmentType vocabulary: {Counter(p[7] for p in P).most_common()}")
    A(f"  Status vocabulary: {Counter(p[10] for p in P).most_common()}")
    A(f"  distinct page Names: {len({p[5] for p in P})}")
    A(f"  distinct form Numbers: {len({p[8] for p in P if p[8]})}")
    A(f"  rows with a non-empty Condition: {sum(1 for p in P if p[11])}")
    A("")
    A(f"PROBLEMS: {len(PR)}")
    for p in PR[:40]:
        A("  " + p)
    (c.OUT / "doc_stc_report.txt").write_text("\n".join(L), encoding="utf-8")
    print("\n".join(L))


if __name__ == "__main__":
    main()
