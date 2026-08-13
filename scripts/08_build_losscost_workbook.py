import os, re, json
from collections import defaultdict, Counter
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PAGES = os.path.join(HERE, "lc_pages")
ROOT = r"C:\Projects\Recursive_Harness_2.0"
OUTX = os.path.join(ROOT, "GL_LossCost_to_ERC.xlsx")

CIRC_RE = re.compile(r"\b(LI-[A-Z]{2,3}-\d{4}-\d{3})\b")
FILING_RE = re.compile(r"\b(GL-\d{4}-[A-Z0-9]{3,8})\b")
NOTICE_RE = re.compile(r"\b(GL-[A-Z]{2}-\d{4}-LC-\d{3})\b")
DATED_CIRC_RE = re.compile(r"(LI-[A-Z]{2,3}-\d{4}-\d{3})\s*\((\d{2}/\d{2}/\d{4})\)")

def dkey(d):
    if not d:
        return "9999-99-99"
    mm, dd, yy = str(d).split("/")
    return f"{yy}-{mm}-{dd}"

# ---------- 1. parse notices ----------
pdfs = []
for fn in sorted(os.listdir(PAGES)):
    if not fn.endswith(".txt"):
        continue
    pdf = fn[:-4] + ".pdf"
    txt = open(os.path.join(PAGES, fn), encoding="utf-8").read()
    head = txt[:4000]
    st = pdf.split("-")[1]
    m = NOTICE_RE.search(head)
    notice = m.group(1) if m else pdf.replace("-C.pdf", "")
    ref = head
    ix = head.upper().find("REFERENCE INFORMATION")
    if ix >= 0:
        ref = head[ix:ix + 1500]
    circs = list(dict.fromkeys(CIRC_RE.findall(ref)))
    # a circular number contains a filing-shaped substring ("LI-GL-2020-100" -> "GL-2020-100"); drop those
    filings = [f for f in dict.fromkeys(FILING_RE.findall(ref))
               if not any(f in c for c in circs)]
    cdates = dict(DATED_CIRC_RE.findall(ref))
    pdfs.append(dict(pdf=pdf, st=st, notice=notice, circs=circs, filings=filings,
                     cdates=cdates, empty=len(txt.strip()) < 40))

# ---------- 2. ERC index ----------
wb = openpyxl.load_workbook(os.path.join(ROOT, "GL_ERC_Edition_Hierarchy.xlsx"), read_only=True)
ws = wb["ERC Circulars"]
rows = list(ws.iter_rows(values_only=True))
h = {k: i for i, k in enumerate(rows[0])}

by_circ = defaultdict(list)      # (ST, circular) -> recs
by_filing = defaultdict(list)    # (ST, filing)   -> recs
circ_meta = {}
for r in rows[1:]:
    st, pkg, ed, cw = r[h["ST"]], r[h["ERC package"]], r[h["Edition date"]], r[h["CW parent"]]
    circ, typ = r[h["Circular"]], r[h["Type"]]
    filing = str(r[h["Filing reference"]] or "")
    desc = r[h["Circular description"]]
    rec = (ed, pkg, cw, typ, filing, desc)
    if circ:
        by_circ[(st, circ)].append(rec)
        if circ not in circ_meta:
            circ_meta[circ] = (desc, typ)
    for f in re.split(r"[,\s]+", filing):
        if FILING_RE.fullmatch(f.strip()):
            by_filing[(st, f.strip())].append(rec)

# every ERC edition per state, for proximity dating
ws2 = wb["Edition Hierarchy"]
er = list(ws2.iter_rows(values_only=True))
h2 = {k: i for i, k in enumerate(er[0])}
eds_by_state = defaultdict(list)
edition_cw = {}
for r in er[1:]:
    st, pkg, ed, cw = r[h2["ST"]], r[h2["ERC package"]], r[h2["Edition date"]], r[h2["CW parent package"]]
    if not pkg:
        continue
    eds_by_state[st].append((dkey(ed), ed, pkg, cw))
    edition_cw[pkg] = (ed, cw)
for st in eds_by_state:
    eds_by_state[st].sort()

# ---------- 3. match ----------
out = []
for p in pdfs:
    st, hits, key_used, method = p["st"], [], None, None
    for c in p["circs"]:
        for scope in (st, "CW"):
            recs = by_circ.get((scope, c))
            if recs:
                hits.extend(recs)
                if key_used is None:
                    key_used, method = c, "Circular reference"
    if not hits:
        for f in p["filings"]:
            for scope in (st, "CW"):
                recs = by_filing.get((scope, f))
                if recs:
                    hits.extend(recs)
                    if key_used is None:
                        key_used, method = f, "Filing reference"
                    break

    pkg = ed = cw = all_eds = ""
    conf = "Unmatched"
    if hits:
        state_hits = [x for x in hits if x[1] and x[1].split()[1] == st]
        pool = sorted(set(state_hits or hits), key=lambda x: (dkey(x[0]), x[1]))
        ed, pkg, cw = pool[0][0], pool[0][1], pool[0][2]
        all_eds = ", ".join(sorted({x[1] for x in pool}, key=lambda k: k))
        conf = "High" if method == "Circular reference" and state_hits else \
               ("Medium" if state_hits else "Low")
    else:
        # tier 3 — earliest ERC edition for this state effective on/after the circular date
        cd = next(iter(p["cdates"].values()), None)
        if cd and eds_by_state.get(st):
            cand = [e for e in eds_by_state[st] if e[0] >= dkey(cd)]
            pick = cand[0] if cand else None
            if pick:
                ed, pkg, cw = pick[1], pick[2], pick[3]
                method, key_used, conf = "Effective-date proximity", cd, "Low"
            else:
                method = "No ERC edition on or after the circular date"
        else:
            method = method or ("No reference found on page 1" if p["empty"] or not p["circs"]
                                else "Reference not present in the ERC corpus")

    desc, ctype = "", ""
    for c in p["circs"]:
        if c in circ_meta:
            desc, ctype = circ_meta[c]
            break

    out.append([
        p["pdf"], st, p["notice"],
        "; ".join(p["circs"]),
        "; ".join(f"{k} ({v})" for k, v in p["cdates"].items()),
        "; ".join(p["filings"]),
        ctype, desc,
        method or "Reference not present in the ERC corpus",
        key_used or "", conf, pkg, ed, cw, all_eds,
        "Page 1 could not be extracted \u2014 PDF is truncated" if p["empty"] else "",
    ])

out.sort(key=lambda r: (r[1], r[0]))

# ---------- 4. reverse view: ERC edition -> loss cost notices ----------
rev = defaultdict(list)
for r in out:
    if r[11]:
        rev[r[11]].append(r[2])
rev_rows = []
for st in sorted(eds_by_state):
    for _, ed, pkg, cw in eds_by_state[st]:
        n = rev.get(pkg, [])
        rev_rows.append([st, pkg, ed, cw or "", len(n), ", ".join(sorted(n))])

# ---------- 5. write workbook ----------
HDR = PatternFill("solid", fgColor="1F3864")
HF = Font(color="FFFFFF", bold=True)
CONF_FILL = {"High": PatternFill("solid", fgColor="C6EFCE"),
             "Medium": PatternFill("solid", fgColor="FFEB9C"),
             "Low": PatternFill("solid", fgColor="FFD9B3"),
             "Unmatched": PatternFill("solid", fgColor="FFC7CE")}

nb = openpyxl.Workbook()

cnt = Counter(r[10] for r in out)
mth = Counter(r[8] for r in out)
rm = nb.active
rm.title = "Read Me"
readme = [
    ["GL Loss Cost Notices \u2192 ERC Editions", ""],
    ["", ""],
    ["Purpose", "Match every General Liability Loss Costs notice PDF to the ISO ERC edition(s) that carry it, "
                "using the same circular-reference method already applied to the Rules manuals."],
    ["Sources", ""],
    ["  Loss cost PDFs", r"Commercial Line Manuals\GL\LossCosts  (472 files)"],
    ["  ERC circular index", r"GL_ERC_Edition_Hierarchy.xlsx  \u2014 sheets 'ERC Circulars' and 'Edition Hierarchy'"],
    ["", ""],
    ["Method (in order; first tier that hits wins)", ""],
    ["  1. Circular reference", "The 'Circular Reference(s):' line on page 1 of the notice (e.g. LI-GL-2022-325) is "
                               "looked up in the ERC circular index for the same state, then countrywide. The matched "
                               "edition is the EARLIEST state ERC edition carrying that circular \u2014 i.e. the edition "
                               "the notice was first implemented into. Confidence: High."],
    ["  2. Filing reference", "If the circular is not in the ERC corpus, the 'Filing Reference(s):' line (e.g. GL-2022-BGL1) "
                             "is looked up instead. Confidence: Medium (state hit) / Low (countrywide only)."],
    ["  3. Effective-date proximity", "If neither key is in the corpus, the notice is assigned to the earliest ERC edition "
                                     "for that state effective ON OR AFTER the circular's issue date. This is dating by "
                                     "proximity, not by citation. Confidence: Low \u2014 verify before relying on it."],
    ["", ""],
    ["Results", ""],
] + [[f"  {k}", v] for k, v in cnt.most_common()] + [
    ["  TOTAL", len(out)],
    ["", ""],
    ["Match method breakdown", ""],
] + [[f"  {k}", v] for k, v in mth.most_common()] + [
    ["", ""],
    ["Known limits", ""],
    ["  Truncated file", "GL-MI-2027-LC-003-C.pdf has no readable page 1 (damaged PDF) and cannot be matched. Re-download."],
    ["  Corpus boundary", "The ERC corpus spans editions 12/01/2020 \u2013 04/01/2027. Notices citing 2019\u20132020 circulars "
                          "predate the earliest ERC edition, and some mid-2026 state loss cost circulars postdate the "
                          "latest edition downloaded for that state. Those are the tier-3 rows."],
    ["  Earliest-edition rule", "'Matched ERC Edition' is the first edition carrying the circular. Column "
                                "'All ERC Editions Carrying This Circular' shows every edition it remains live in."],
]
for r in readme:
    rm.append(r)
rm["A1"].font = Font(bold=True, size=14)
for c in ("A4", "A8", "A13", "A" + str(readme.index(["Known limits", ""]) + 1)):
    rm[c].font = Font(bold=True)
rm.column_dimensions["A"].width = 32
rm.column_dimensions["B"].width = 120
for row in rm.iter_rows():
    for c in row:
        c.alignment = Alignment(vertical="top", wrap_text=True)

def sheet(name, headers, data, widths, conf_col=None):
    s = nb.create_sheet(name)
    s.append(headers)
    for c in s[1]:
        c.fill, c.font = HDR, HF
        c.alignment = Alignment(vertical="center", wrap_text=True)
    for r in data:
        s.append(r)
    for i, w in enumerate(widths, 1):
        s.column_dimensions[get_column_letter(i)].width = w
    s.freeze_panes = "C2"
    s.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"
    if conf_col:
        for r in range(2, len(data) + 2):
            v = s.cell(r, conf_col).value
            if v in CONF_FILL:
                s.cell(r, conf_col).fill = CONF_FILL[v]
    return s

sheet("LC to ERC",
      ["Loss Cost PDF", "ST", "Notice Number", "Circular Reference(s)", "Circular Issue Date",
       "Filing Reference(s)", "Circular Type", "Circular Description", "Match Method", "Key Used",
       "Confidence", "Matched ERC Edition", "ERC Edition Date", "CW Parent",
       "All ERC Editions Carrying This Circular", "Notes"],
      out,
      [26, 5, 22, 22, 26, 24, 20, 60, 24, 18, 12, 22, 14, 22, 70, 40],
      conf_col=11)

sheet("ERC to LC",
      ["ST", "ERC Edition", "Edition Date", "CW Parent", "# Loss Cost Notices", "Loss Cost Notices"],
      rev_rows, [5, 22, 14, 24, 10, 90])

gaps = [r for r in out if r[10] in ("Unmatched", "Low")]
sheet("Gaps",
      ["Loss Cost PDF", "ST", "Notice Number", "Circular Reference(s)", "Circular Issue Date",
       "Filing Reference(s)", "Match Method", "Confidence", "Matched ERC Edition", "ERC Edition Date", "Notes"],
      [[r[0], r[1], r[2], r[3], r[4], r[5], r[8], r[10], r[11], r[12], r[15]] for r in gaps],
      [26, 5, 22, 22, 26, 24, 30, 12, 22, 14, 40], conf_col=8)

nb.save(OUTX)
print("wrote", OUTX)
print(cnt, len(out))
print("ERC editions with >=1 loss cost notice:", sum(1 for r in rev_rows if r[4]), "of", len(rev_rows))
