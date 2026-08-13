"""Build the ISO Circular Expert knowledge base from the verified analyses."""
import json, os, re, sys
from collections import defaultdict, Counter
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = r"C:\Projects\Recursive_Harness_2.0"
KB = os.path.join(ROOT, "Agentic", "iso-circular-expert", "knowledge")
os.makedirs(KB, exist_ok=True)

lc = json.load(open(os.path.join(HERE, "lc_analysis2.json")))
mode = json.load(open(os.path.join(HERE, "lc_extractor.json")))
lcmatch = {m["pdf"]: m for m in json.load(open(os.path.join(HERE, "lc_match.json")))["out"]}
terr = json.load(open(os.path.join(HERE, "territory_scan.json")))
ds = json.load(open(os.path.join(ROOT, "docs", "rating-engine", "dataset.json"), encoding="utf-8"))

import openpyxl
wb = openpyxl.load_workbook(os.path.join(ROOT, "GL_ERC_Edition_Hierarchy.xlsx"), read_only=True)

# ---------------------------------------------------------------- circulars
ws = wb["ERC Circulars"]
rows = list(ws.iter_rows(values_only=True))
h = {k: i for i, k in enumerate(rows[0])}
circ = {}
for r in rows[1:]:
    c = r[h["Circular"]]
    if not c:
        continue
    e = circ.setdefault(c, {"circular": c, "description": r[h["Circular description"]],
                            "type": r[h["Type"]], "filings": set(), "editions": [],
                            "states": set(), "effective_dates": set()})
    if r[h["Filing reference"]]:
        for f in re.split(r"[,\s]+", str(r[h["Filing reference"]])):
            if re.fullmatch(r"GL-\d{4}-[A-Z0-9]{3,8}", f.strip()):
                e["filings"].add(f.strip())
    e["editions"].append({"st": r[h["ST"]], "erc": r[h["ERC package"]],
                          "edition_date": r[h["Edition date"]], "cw_parent": r[h["CW parent"]]})
    e["states"].add(r[h["ST"]])
    if r[h["Eff. date (as stated in THIS package)"]]:
        e["effective_dates"].add(str(r[h["Eff. date (as stated in THIS package)"]]))
    if not e["description"] and r[h["Circular description"]]:
        e["description"] = r[h["Circular description"]]
for e in circ.values():
    e["filings"] = sorted(e["filings"])
    e["states"] = sorted(e["states"])
    e["effective_dates"] = sorted(e["effective_dates"])

# rules-notice -> circular, from the ERC workbook's own PDF match sheet
ws2 = wb["Rules PDFs"]
r2 = list(ws2.iter_rows(values_only=True))
h2 = {k: i for i, k in enumerate(r2[0])}
rules_notices = {}
for r in r2[1:]:
    f = r[h2["PDF file"]]
    if not f:
        continue
    rules_notices[f] = {
        "file": f, "kind": "RU", "st": r[h2["File ST"]],
        "notice": r[h2["Notice number"]],
        "pages": r[h2["Pages"]],
        "circulars": [x.strip() for x in re.split(r"[,\s]+", str(r[h2["Circular reference(s)"]] or "")) if x.strip()],
        "filings": [x.strip() for x in re.split(r"[,\s]+", str(r[h2["Filing reference(s)"]] or "")) if x.strip()],
        "effective_date": r[h2["Manual effective date"]],
        "dating_basis": r[h2["Dating basis"]],
        "edition_marker": r[h2["Newest page edition marker"]],
        "erc_edition": r[h2["Matched ERC package"]],
        "erc_edition_date": r[h2["Matched edition date"]],
        "cw_parent": r[h2["CW parent of match"]],
        "date_confidence": r[h2["Confidence"]],
    }

# ---------------------------------------------------------------- loss cost notices
lc_notices = {}
for f, d in lc.items():
    m = lcmatch.get(f, {})
    lc_notices[f] = {
        "file": f, "kind": "LC", "st": d["st"], "notice": f[:-6],
        "year": d["year"], "pages": d["pages"],
        "circulars": [c for c in (m.get("circs") or "").split("; ") if c],
        "filings": [x for x in (m.get("filings") or "").split("; ") if x],
        "circular_issue_date": m.get("circ_dates") or None,
        "circular_description": m.get("circ_desc") or None,
        "erc_edition": m.get("pkg") or None,
        "erc_edition_date": m.get("ed") or None,
        "cw_parent": m.get("cw") or None,
        "date_confidence": m.get("conf"),
        "match_method": m.get("method"),
        "territories": d["territories"],
        "lc_pages": 8 * d["n_territories"] + 1,
        "class_codes": d["lc_classes_total"],
        "elp_classes": d["elp_classes"],
        "extractor": "pypdf" if mode.get(f, {}).get("mode") == "PYPDF_FALLBACK" else "pdftotext",
    }

# ---------------------------------------------------------------- jurisdictions
lcby = defaultdict(list)
for d in lc.values():
    lcby[d["st"]].append(d)
lclatest = {st: max(v, key=lambda d: d["file"]) for st, v in lcby.items()}
tby = defaultdict(list)
for f, r in terr.items():
    if "error" in r or "(" in f or r.get("st") == "MU":
        continue
    tby[r["st"]].append(r)
tlatest = {st: max(v, key=lambda r: r["file"]) for st, v in tby.items()}

TD = ds["territory_definitions"]["states"]
LCS = ds["loss_costs"]["states"]
DSTATES = ds["states"]

juris = {}
for st in sorted(lclatest):
    d = lclatest[st]
    t = tlatest.get(st, {})
    s = DSTATES.get(st, {})
    juris[st] = {
        "st": st,
        "latest_rules_notice": t.get("file", "")[:-6] or None,
        "latest_losscost_notice": d["file"][:-6],
        "rules_notices_held": len([1 for f in rules_notices if f.split("-")[1] == st]),
        "losscost_notices_held": len(lcby[st]),
        "territory": {
            "scheme": TD[st]["scheme"],
            "count": TD[st]["territories_lc"],
            "domain": d["territories"],
            "cgt_pages": TD[st]["cgt_pages"],
            "zip_rows": TD[st]["zip_rows"],
            "place_rows": TD[st]["place_rows"],
            "cross_corpus_match": TD[st]["match"],
            "territorial_sublines": ["334", "332"],
            "statewide_sublines": ["335", "336", "350"],
        },
        "rates": {
            "vintage": LCS[st]["vintage"],
            "class_codes": LCS[st]["classes"],
            "elp_classes": LCS[st]["elp_classes"],
            "ocp_loss_costs_published": LCS[st]["ocp_loss_costs"],
            "extractor": LCS[st]["extractor"],
        },
        "variables": {
            "payroll": s.get("payroll"),
            "liquor_grades": s.get("liquor_grades"),
            "ilf_table_count": s.get("ilf_table_count"),
            "ilf_by_subline": s.get("ilf_by_subline"),
            "basic_limits": s.get("basic_limits"),
            "stopgap": s.get("stopgap"),
            "elp_referenced": s.get("elp"),
        },
        "deviations": s.get("kinds"),
        "additional_rules": s.get("additional_rules"),
        "deviation_count": s.get("deviation_count"),
    }

# ---------------------------------------------------------------- write
def dump(name, obj):
    p = os.path.join(KB, name)
    json.dump(obj, open(p, "w", encoding="utf-8"), indent=1, default=str)
    print(f"  {name:24} {os.path.getsize(p)/1024:8.0f} KB")

print("knowledge base:")
dump("circulars.json", circ)
dump("notices.json", {"rules": rules_notices, "losscosts": lc_notices})
dump("jurisdictions.json", juris)
print("\ncounts: circulars", len(circ), "| rules notices", len(rules_notices),
      "| lc notices", len(lc_notices), "| jurisdictions", len(juris))
